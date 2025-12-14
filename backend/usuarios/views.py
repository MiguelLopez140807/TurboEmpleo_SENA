from .models import Postulacion, Notificacion, Vacante, Empresa
from .serializers import PostulacionSerializer, NotificacionSerializer
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db import IntegrityError
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import io
from django.http import HttpResponse

# Imports para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Imports para PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import tempfile
import os

class PostulacionViewSet(viewsets.ModelViewSet):
    queryset = Postulacion.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            from .serializers import PostulacionWriteSerializer
            return PostulacionWriteSerializer
        return PostulacionSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError as e:
            # Capturar errores de unique_together de la base de datos
            if 'pos_aspirante_fk' in str(e) and 'pos_vacante_fk' in str(e):
                return Response(
                    {"detail": "Ya te has postulado a esta vacante anteriormente. No puedes postularte dos veces a la misma oferta laboral."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            return Response(
                {"detail": "Error al procesar la postulación."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def get_queryset(self):
        queryset = Postulacion.objects.all()
        
        # Filtrar por aspirante
        aspirante = self.request.query_params.get('pos_aspirante_fk', None)
        if aspirante is not None:
            queryset = queryset.filter(pos_aspirante_fk=aspirante)
        
        # Filtrar por vacante
        vacante = self.request.query_params.get('pos_vacante_fk', None)
        if vacante is not None:
            queryset = queryset.filter(pos_vacante_fk=vacante)
        
        # Filtrar por estado
        estado = self.request.query_params.get('pos_estado', None)
        if estado is not None:
            queryset = queryset.filter(pos_estado=estado)
        
        # Ordenar por fecha de postulación (más recientes primero)
        queryset = queryset.order_by('-pos_fechaPostulacion')
        
        return queryset
    
    # ⚡ ENDPOINTS MODO TURBO
    @action(detail=False, methods=['get'])
    def turbo(self, request):
        """
        Endpoint: GET /api/postulaciones/turbo/
        Devuelve solo postulaciones en modo turbo
        """
        postulaciones_turbo = self.queryset.filter(pos_es_turbo=True)
        
        # Filtrar por aspirante si se especifica
        aspirante = request.query_params.get('aspirante', None)
        if aspirante:
            postulaciones_turbo = postulaciones_turbo.filter(pos_aspirante_fk=aspirante)
        
        serializer = self.get_serializer(postulaciones_turbo, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def turbo_pendientes(self, request):
        """
        Endpoint: GET /api/postulaciones/turbo_pendientes/?empresa=<id>
        Devuelve postulaciones turbo pendientes de respuesta para una empresa
        """
        from django.utils import timezone
        
        empresa = request.query_params.get('empresa', None)
        if not empresa:
            return Response(
                {'error': 'Parámetro empresa es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        postulaciones_pendientes = self.queryset.filter(
            pos_es_turbo=True,
            pos_estado='Pendiente',
            pos_vacante_fk__va_idEmpresa_fk=empresa,
            pos_fecha_limite_respuesta__gt=timezone.now()
        ).order_by('pos_fecha_limite_respuesta')
        
        serializer = self.get_serializer(postulaciones_pendientes, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def responder_turbo(self, request, pk=None):
        """
        Endpoint: POST /api/postulaciones/<id>/responder_turbo/
        Marca una postulación turbo como respondida y actualiza el score de la empresa
        Body: { "nuevo_estado": "Aceptada" | "Rechazada" }
        """
        from django.utils import timezone
        
        postulacion = self.get_object()
        
        if not postulacion.pos_es_turbo:
            return Response(
                {'error': 'Esta postulación no es turbo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        nuevo_estado = request.data.get('nuevo_estado')
        if nuevo_estado not in ['Aceptada', 'Rechazada']:
            return Response(
                {'error': 'Estado debe ser Aceptada o Rechazada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Actualizar estado
        postulacion.pos_estado = nuevo_estado
        
        # Verificar si respondió a tiempo
        ahora = timezone.now()
        if ahora <= postulacion.pos_fecha_limite_respuesta:
            postulacion.pos_respondida_a_tiempo = True
        
        postulacion.save()
        
        serializer = self.get_serializer(postulacion)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def activar_turbo_aspirante(self, request, pk=None):
        """
        Endpoint: POST /api/postulaciones/<id>/activar_turbo_aspirante/
        Permite al aspirante activar modo turbo en una postulación existente
        """
        from django.utils import timezone
        from datetime import timedelta
        
        postulacion = self.get_object()
        aspirante = postulacion.pos_aspirante_fk
        vacante = postulacion.pos_vacante_fk
        
        # Validaciones
        if postulacion.pos_es_turbo:
            return Response(
                {'error': 'Esta postulación ya está en modo turbo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if postulacion.pos_estado != 'Pendiente':
            return Response(
                {'error': 'Solo puedes activar turbo en postulaciones pendientes'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if aspirante.asp_creditos_turbo_disponibles <= 0:
            return Response(
                {'error': f'No tienes créditos turbo disponibles. Créditos restantes: {aspirante.asp_creditos_turbo_disponibles}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Activar modo turbo
        postulacion.pos_es_turbo = True
        postulacion.pos_turbo_solicitado_por_aspirante = True
        postulacion.pos_creditos_turbo_usados = 1
        
        # Determinar horas de respuesta
        if vacante.va_modo_turbo:
            horas_limite = vacante.va_tiempo_respuesta_horas
        else:
            horas_limite = 48  # Default para turbo del aspirante
        
        # Establecer fecha límite
        postulacion.pos_fecha_limite_respuesta = timezone.now() + timedelta(hours=horas_limite)
        postulacion.save()
        
        # Descontar crédito del aspirante
        aspirante.asp_creditos_turbo_disponibles -= 1
        aspirante.asp_creditos_turbo_usados += 1
        aspirante.save()
        
        # Incrementar contador de empresa
        empresa = vacante.va_idEmpresa_fk
        empresa.em_total_postulaciones_turbo += 1
        empresa.save()
        
        # Crear notificación para el aspirante
        from .models import Notificacion
        usuario_aspirante = aspirante.asp_usuario_fk
        Notificacion.objects.create(
            not_usuario_fk=usuario_aspirante,
            not_contenido=f"⚡ Modo Turbo activado para '{vacante.va_titulo}'. Respuesta prioritaria en {horas_limite} horas. Créditos restantes: {aspirante.asp_creditos_turbo_disponibles}",
            not_estado='No leída'
        )
        
        # Crear notificación para la empresa
        usuario_empresa = empresa.em_usuario_fk
        Notificacion.objects.create(
            not_usuario_fk=usuario_empresa,
            not_contenido=f"⚡ {aspirante.asp_nombre} {aspirante.asp_apellido} solicita respuesta TURBO para '{vacante.va_titulo}'",
            not_estado='No leída'
        )
        
        serializer = self.get_serializer(postulacion)
        return Response({
            'message': 'Modo turbo activado exitosamente',
            'creditos_restantes': aspirante.asp_creditos_turbo_disponibles,
            'postulacion': serializer.data
        })
from .models import ExperienciaLaboral, ExperienciaEscolar
from rest_framework import viewsets
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import UsuarioSerializer, AspiranteSerializer, EmpresaSerializer, VacanteSerializer, UsuarioRegistroSerializer, ExperienciaLaboralSerializer, ExperienciaEscolarSerializer
# ViewSet para ExperienciaLaboral
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated

class ExperienciaLaboralViewSet(viewsets.ModelViewSet):
    queryset = ExperienciaLaboral.objects.all()
    serializer_class = ExperienciaLaboralSerializer
    permission_classes = [IsAuthenticated]

# ViewSet para ExperienciaEscolar
class ExperienciaEscolarViewSet(viewsets.ModelViewSet):
    queryset = ExperienciaEscolar.objects.all()
    serializer_class = ExperienciaEscolarSerializer
    permission_classes = [IsAuthenticated]
from .models import Usuarios, Aspirante, Empresa, Vacante

from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import MyTokenObtainPairSerializer
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuarios.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=True, methods=['patch'])
    def toggle_status(self, request, pk=None):
        """
        Endpoint para cambiar el estado activo/inactivo de un usuario
        """
        try:
            usuario = self.get_object()
            nuevo_estado = request.data.get('is_active', not usuario.is_active)
            usuario.is_active = nuevo_estado
            usuario.save()
            
            return Response({
                'message': f'Usuario {"activado" if nuevo_estado else "desactivado"} exitosamente',
                'is_active': usuario.is_active
            })
        except Exception as e:
            return Response(
                {'error': 'Error al cambiar estado del usuario'},
                status=400
            )

from rest_framework import filters

class AspiranteViewSet(viewsets.ModelViewSet):
    queryset = Aspirante.objects.all()
    serializer_class = AspiranteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['asp_usuario_fk__user_nombre']

class EmpresaViewSet(viewsets.ModelViewSet):
    queryset = Empresa.objects.all()
    serializer_class = EmpresaSerializer
    permission_classes = [IsAuthenticated]
    
    def partial_update(self, request, *args, **kwargs):
        """Sobrescribir para manejar cambios de estado y logging temporal"""
        print(f"Datos recibidos para empresa PATCH: {request.data}")
        
        # Si se está actualizando el estado, actualizar el usuario asociado
        if 'em_estado' in request.data:
            instance = self.get_object()
            nuevo_estado = request.data['em_estado']
            is_active = nuevo_estado == 'activo'
            
            # Actualizar el estado del usuario asociado
            if instance.em_usuario_fk:
                instance.em_usuario_fk.is_active = is_active
                instance.em_usuario_fk.save()
                print(f"Estado del usuario actualizado: {is_active}")
        
        try:
            response = super().partial_update(request, *args, **kwargs)
            print(f"Actualización exitosa")
            return response
        except Exception as e:
            print(f"Error en partial_update: {str(e)}")
            # Verificar errores del serializer
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if not serializer.is_valid():
                print(f"Errores de validación del serializer: {serializer.errors}")
            raise e
    
    # ⚡ ENDPOINTS MODO TURBO
    @action(detail=False, methods=['get'])
    def ranking_turbo(self, request):
        """
        Endpoint: GET /api/empresas/ranking_turbo/
        Devuelve ranking de empresas por score turbo (mejores primero)
        """
        empresas_con_turbo = self.queryset.filter(
            em_total_postulaciones_turbo__gt=0
        ).order_by('-em_score_turbo', '-em_respuestas_a_tiempo')
        
        serializer = self.get_serializer(empresas_con_turbo, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def estadisticas_turbo(self, request, pk=None):
        """
        Endpoint: GET /api/empresas/<id>/estadisticas_turbo/
        Devuelve estadísticas detalladas de modo turbo para una empresa
        """
        empresa = self.get_object()
        
        tasa_respuesta = 0
        if empresa.em_total_postulaciones_turbo > 0:
            tasa_respuesta = (empresa.em_respuestas_a_tiempo / empresa.em_total_postulaciones_turbo) * 100
        
        estadisticas = {
            'em_score_turbo': empresa.em_score_turbo,
            'em_score_turbo_calculado': empresa.calcular_score_turbo(),
            'em_total_postulaciones_turbo': empresa.em_total_postulaciones_turbo,
            'em_respuestas_a_tiempo': empresa.em_respuestas_a_tiempo,
            'respuestas_tarde': empresa.em_total_postulaciones_turbo - empresa.em_respuestas_a_tiempo,
            'tasa_respuesta_porcentaje': round(tasa_respuesta, 1),
            'vacantes_turbo_activas': Vacante.objects.filter(
                va_idEmpresa_fk=empresa,
                va_modo_turbo=True,
                va_estado='Activa'
            ).count()
        }
        
        return Response(estadisticas)


class VacanteViewSet(viewsets.ModelViewSet):
    queryset = Vacante.objects.all()
    serializer_class = VacanteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['va_titulo', 'va_descripcion', 'va_requisitos']
    
    def get_serializer_class(self):
        # Usar VacanteWriteSerializer para crear/actualizar
        if self.action in ['create', 'update', 'partial_update']:
            from .serializers import VacanteWriteSerializer
            return VacanteWriteSerializer
        # Usar VacanteSerializer para listar/obtener (con datos de empresa)
        return VacanteSerializer
    
    def get_queryset(self):
        queryset = Vacante.objects.all()
        
        # Filtrar por empresa
        empresa = self.request.query_params.get('empresa', None)
        if empresa is not None:
            queryset = queryset.filter(va_idEmpresa_fk=empresa)
        
        # Filtrar por ubicación
        ubicacion = self.request.query_params.get('ubicacion', None)
        if ubicacion is not None:
            queryset = queryset.filter(va_ubicacion__icontains=ubicacion)
        
        # Filtrar por tipo de empleo
        tipo_empleo = self.request.query_params.get('tipo_empleo', None)
        if tipo_empleo is not None:
            queryset = queryset.filter(va_tipo_empleo=tipo_empleo)
        
        # Filtrar por estado (activa/inactiva)
        estado = self.request.query_params.get('estado', None)
        if estado is not None:
            queryset = queryset.filter(va_estado=estado)
        
        # Ordenar por fecha de publicación (más recientes primero)
        queryset = queryset.order_by('-va_fecha_publicacion')
        
        return queryset
    
    # ⚡ ENDPOINTS MODO TURBO
    @action(detail=False, methods=['get'])
    def turbo(self, request):
        """
        Endpoint: GET /api/vacantes/turbo/
        Devuelve solo vacantes con modo turbo activado
        """
        vacantes_turbo = self.queryset.filter(va_modo_turbo=True, va_estado='Activa')
        serializer = self.get_serializer(vacantes_turbo, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def turbo_recomendadas(self, request):
        """
        Endpoint: GET /api/vacantes/turbo_recomendadas/
        Devuelve vacantes turbo de empresas con mejor score
        """
        from django.db.models import F
        
        vacantes_turbo = self.queryset.filter(
            va_modo_turbo=True, 
            va_estado='Activa'
        ).select_related('va_idEmpresa_fk').order_by(
            '-va_idEmpresa_fk__em_score_turbo',
            '-va_fecha_publicacion'
        )
        
        serializer = self.get_serializer(vacantes_turbo, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def import_vacantes(self, request):
        """Importar vacantes desde archivo CSV"""
        try:
            csv_file = request.FILES.get('file')
            if not csv_file:
                return Response(
                    {'error': 'No se proporcionó ningún archivo'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not csv_file.name.endswith('.csv'):
                return Response(
                    {'error': 'El archivo debe ser de formato CSV'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Leer archivo CSV
            decoded_file = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded_file))
            
            vacantes_creadas = 0
            errores = []
            
            for row_num, row in enumerate(csv_reader, 1):
                try:
                    # Validar empresa existe
                    empresa = Empresa.objects.get(id=row['empresa_id'])
                    
                    # Preparar datos de la vacante
                    datos_vacante = {
                        'va_titulo': row['nombre'],
                        'va_descripcion': row['descripcion'],
                        'va_ubicacion': row['ubicacion'],
                        'va_salario': float(row['salario']) if row['salario'] else 0,
                        'va_requisitos': f"Experiencia: {row['experiencia']}",
                        'va_tipo_empleo': row['tipo_contrato'],
                        'va_idEmpresa_fk': empresa,
                        'va_estado': 'Activa'
                    }
                    
                    # Agregar campos opcionales si existen
                    if row.get('habilidades'):
                        datos_vacante['va_habilidades'] = row['habilidades']
                    
                    if row.get('beneficios'):
                        datos_vacante['va_beneficios'] = row['beneficios']
                    
                    # Crear vacante
                    vacante = Vacante.objects.create(**datos_vacante)
                    vacantes_creadas += 1
                    
                except Empresa.DoesNotExist:
                    errores.append(f"Fila {row_num}: Empresa con ID {row['empresa_id']} no existe")
                except Exception as e:
                    errores.append(f"Fila {row_num}: {str(e)}")
            
            return Response({
                'message': f'Importación completada. {vacantes_creadas} vacantes creadas.',
                'errores': errores
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Descargar plantilla CSV para importar vacantes"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="plantilla_vacantes.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'empresa_id', 'nombre', 'descripcion', 'ubicacion', 
            'salario', 'experiencia', 'tipo_contrato',
            'habilidades', 'beneficios'
        ])
        
        # Agregar fila de ejemplo
        writer.writerow([
            '1', 'Desarrollador Python', 'Desarrollo de aplicaciones web', 
            'Bogotá', '3500000', 'Junior', 'Indefinido',
            'Python, Django, React', 'Seguro médico, Auxilios'
        ])
        
        return response

class DetalleVacanteViewSet(viewsets.ModelViewSet):
    pass  # Eliminado Detalle_Vacante




class UsuarioRegistroView(APIView):
    def post(self, request):
        serializer = UsuarioRegistroSerializer(data=request.data)
        if serializer.is_valid():
            usuario = serializer.save()
            return Response(
                {"message": "Usuario registrado exitosamente"},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

# ViewSet para Notificaciones
class NotificacionViewSet(viewsets.ModelViewSet):
    queryset = Notificacion.objects.all()
    serializer_class = NotificacionSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filtrar notificaciones del usuario actual"""
        user = self.request.user
        queryset = Notificacion.objects.filter(not_usuario_fk=user).order_by('-not_fecha')
        
        # Filtrar por estado (leída/no leída)
        estado = self.request.query_params.get('not_estado', None)
        if estado is not None:
            queryset = queryset.filter(not_estado=estado)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def marcar_leida(self, request, pk=None):
        """Marcar una notificación como leída"""
        notificacion = self.get_object()
        notificacion.not_estado = 'Leída'
        notificacion.save()
        serializer = self.get_serializer(notificacion)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def marcar_todas_leidas(self, request):
        """Marcar todas las notificaciones del usuario como leídas"""
        user = request.user
        Notificacion.objects.filter(not_usuario_fk=user, not_estado='No leída').update(not_estado='Leída')
        return Response({'message': 'Todas las notificaciones han sido marcadas como leídas'})
    
    @action(detail=False, methods=['get'])
    def no_leidas_count(self, request):
        """Obtener el conteo de notificaciones no leídas"""
        user = request.user
        count = Notificacion.objects.filter(not_usuario_fk=user, not_estado='No leída').count()
        return Response({'count': count})


# Vista para el formulario de contacto
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from django.core.mail import send_mail
from django.conf import settings

@api_view(['POST'])
@permission_classes([AllowAny])
def contacto_view(request):
    """
    Vista para procesar el formulario de contacto.
    Envía un email con los datos del formulario.
    """
    try:
        nombre = request.data.get('nombre', '')
        email = request.data.get('email', '')
        asunto = request.data.get('asunto', '')
        mensaje = request.data.get('mensaje', '')
        
        # Validar que todos los campos estén presentes
        if not all([nombre, email, asunto, mensaje]):
            return Response(
                {'error': 'Todos los campos son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Construir el mensaje del email
        mensaje_email = f"""
Nuevo mensaje de contacto de TurboEmpleo:

Nombre: {nombre}
Email: {email}
Asunto: {asunto}

Mensaje:
{mensaje}
"""
        
        # En desarrollo, imprimir en consola
        print('='*60)
        print('NUEVO MENSAJE DE CONTACTO')
        print('='*60)
        print(mensaje_email)
        print('='*60)
        
        # Intentar enviar email (opcional en desarrollo)
        try:
            # Configurar el email solo si está configurado en settings
            if hasattr(settings, 'EMAIL_HOST') and settings.EMAIL_HOST:
                send_mail(
                    subject=f'Contacto TurboEmpleo: {asunto}',
                    message=mensaje_email,
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@turboempleo.co'),
                    recipient_list=[getattr(settings, 'CONTACT_EMAIL', 'contacto@turboempleo.co')],
                    fail_silently=True,
                )
        except Exception as email_error:
            # Si falla el email, solo lo registramos pero no fallar la petición
            print(f"Advertencia: No se pudo enviar email: {str(email_error)}")
        
        return Response(
            {'message': 'Mensaje enviado con éxito'},
            status=status.HTTP_200_OK
        )
            
    except Exception as e:
        print(f"Error en contacto_view: {str(e)}")
        return Response(
            {'error': 'Error al procesar el mensaje'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# ViewSet para Reportes Parametrizados
class ReportesViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def postulaciones(self, request):
        """Reporte de postulaciones con filtros parametrizados"""
        try:
            # Obtener parámetros de filtrado
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            estado = request.query_params.get('estado')
            empresa_id = request.query_params.get('empresa_id')
            aspirante_id = request.query_params.get('aspirante_id')
            formato = request.query_params.get('formato', 'json')  # json, csv, excel
            
            # Query base
            queryset = Postulacion.objects.select_related(
                'pos_vacante_fk__va_idEmpresa_fk',
                'pos_aspirante_fk'
            )
            
            # Aplicar filtros
            if fecha_inicio:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                queryset = queryset.filter(pos_fechaPostulacion__gte=fecha_inicio_dt)
            
            if fecha_fin:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                queryset = queryset.filter(pos_fechaPostulacion__lte=fecha_fin_dt)
            
            if estado:
                queryset = queryset.filter(pos_estado=estado)
            
            if empresa_id:
                queryset = queryset.filter(pos_vacante_fk__va_idEmpresa_fk__id=empresa_id)
            
            if aspirante_id:
                queryset = queryset.filter(pos_aspirante_fk__id=aspirante_id)
            
            # Estadísticas
            total_postulaciones = queryset.count()
            por_estado = queryset.values('pos_estado').annotate(count=Count('id'))
            por_empresa = queryset.values(
                'pos_vacante_fk__va_idEmpresa_fk__em_nombre'
            ).annotate(count=Count('id'))[:10]
            
            # Preparar datos
            postulaciones_data = []
            for post in queryset:
                postulaciones_data.append({
                    'id': post.id,
                    'fecha': post.pos_fechaPostulacion.strftime('%Y-%m-%d %H:%M'),
                    'estado': post.pos_estado,
                    'aspirante': f"{post.pos_aspirante_fk.asp_nombre} {post.pos_aspirante_fk.asp_apellido}",
                    'empresa': post.pos_vacante_fk.va_idEmpresa_fk.em_nombre,
                    'vacante': post.pos_vacante_fk.va_titulo,
                    'es_turbo': post.pos_es_turbo
                })
            
            resultado = {
                'resumen': {
                    'total_postulaciones': total_postulaciones,
                    'filtros_aplicados': {
                        'fecha_inicio': fecha_inicio,
                        'fecha_fin': fecha_fin,
                        'estado': estado,
                        'empresa_id': empresa_id,
                        'aspirante_id': aspirante_id
                    }
                },
                'estadisticas': {
                    'por_estado': list(por_estado),
                    'por_empresa': list(por_empresa)
                },
                'datos': postulaciones_data
            }
            
            # Exportar según formato
            if formato == 'csv':
                return self._export_csv(postulaciones_data, 'reporte_postulaciones')
            elif formato == 'excel':
                return self._export_excel(postulaciones_data, 'reporte_postulaciones', 'Postulaciones')
            elif formato == 'pdf':
                return self._export_pdf(postulaciones_data, 'reporte_postulaciones', 'Reporte de Postulaciones')
            
            return Response(resultado)
            
        except Exception as e:
            return Response(
                {'error': f'Error generando reporte: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def vacantes(self, request):
        """Reporte de vacantes con filtros parametrizados"""
        try:
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            estado = request.query_params.get('estado')
            empresa_id = request.query_params.get('empresa_id')
            formato = request.query_params.get('formato', 'json')
            
            queryset = Vacante.objects.select_related('va_idEmpresa_fk')
            
            # Aplicar filtros
            if fecha_inicio:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                queryset = queryset.filter(va_fecha_publicacion__gte=fecha_inicio_dt)
            
            if fecha_fin:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                queryset = queryset.filter(va_fecha_publicacion__lte=fecha_fin_dt)
            
            if estado:
                queryset = queryset.filter(va_estado=estado)
            
            if empresa_id:
                queryset = queryset.filter(va_idEmpresa_fk__id=empresa_id)
            
            # Estadísticas
            total_vacantes = queryset.count()
            por_estado = queryset.values('va_estado').annotate(count=Count('id'))
            por_empresa = queryset.values('va_idEmpresa_fk__em_nombre').annotate(count=Count('id'))
            promedio_salario = queryset.aggregate(promedio=Avg('va_salario'))['promedio']
            
            # Datos detallados
            vacantes_data = []
            for vacante in queryset:
                # Contar postulaciones
                postulaciones_count = Postulacion.objects.filter(pos_vacante_fk=vacante).count()
                
                vacantes_data.append({
                    'id': vacante.id,
                    'titulo': vacante.va_titulo,
                    'empresa': vacante.va_idEmpresa_fk.em_nombre,
                    'fecha_publicacion': vacante.va_fecha_publicacion.strftime('%Y-%m-%d %H:%M'),
                    'estado': vacante.va_estado,
                    'salario': float(vacante.va_salario) if vacante.va_salario else None,
                    'ubicacion': vacante.va_ubicacion,
                    'postulaciones': postulaciones_count,
                    'modo_turbo': vacante.va_modo_turbo
                })
            
            resultado = {
                'resumen': {
                    'total_vacantes': total_vacantes,
                    'promedio_salario': promedio_salario,
                    'filtros_aplicados': {
                        'fecha_inicio': fecha_inicio,
                        'fecha_fin': fecha_fin,
                        'estado': estado,
                        'empresa_id': empresa_id
                    }
                },
                'estadisticas': {
                    'por_estado': list(por_estado),
                    'por_empresa': list(por_empresa)
                },
                'datos': vacantes_data
            }
            
            if formato == 'csv':
                return self._export_csv(vacantes_data, 'reporte_vacantes')
            elif formato == 'excel':
                return self._export_excel(vacantes_data, 'reporte_vacantes', 'Vacantes')
            elif formato == 'pdf':
                return self._export_pdf(vacantes_data, 'reporte_vacantes', 'Reporte de Vacantes')
            
            return Response(resultado)
            
        except Exception as e:
            return Response(
                {'error': f'Error generando reporte: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def usuarios(self, request):
        """Reporte de usuarios (aspirantes y empresas)"""
        try:
            fecha_inicio = request.query_params.get('fecha_inicio')
            fecha_fin = request.query_params.get('fecha_fin')
            tipo_usuario = request.query_params.get('tipo')  # 'aspirante' o 'empresa'
            formato = request.query_params.get('formato', 'json')
            
            resultado = {
                'resumen': {
                    'filtros_aplicados': {
                        'fecha_inicio': fecha_inicio,
                        'fecha_fin': fecha_fin,
                        'tipo_usuario': tipo_usuario
                    }
                },
                'datos': {
                    'aspirantes': [],
                    'empresas': []
                }
            }
            
            # Reporte de aspirantes
            if not tipo_usuario or tipo_usuario == 'aspirante':
                from .models import Aspirante
                aspirantes_query = Aspirante.objects.all()
                
                # Aplicar filtros de fecha si existen los campos
                try:
                    if fecha_inicio and hasattr(Aspirante._meta.get_field('asp_fecha_registro'), 'name'):
                        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                        aspirantes_query = aspirantes_query.filter(asp_fecha_registro__gte=fecha_inicio_dt)
                    
                    if fecha_fin and hasattr(Aspirante._meta.get_field('asp_fecha_registro'), 'name'):
                        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                        aspirantes_query = aspirantes_query.filter(asp_fecha_registro__lte=fecha_fin_dt)
                except:
                    # Si no existe el campo de fecha, continuar sin filtro
                    pass
                
                for aspirante in aspirantes_query:
                    try:
                        postulaciones_count = Postulacion.objects.filter(pos_aspirante_fk=aspirante).count()
                        
                        # Obtener fecha de registro de forma segura
                        fecha_registro = None
                        if hasattr(aspirante, 'asp_fecha_registro') and aspirante.asp_fecha_registro:
                            fecha_registro = aspirante.asp_fecha_registro.strftime('%Y-%m-%d %H:%M')
                        
                        # Obtener campos de forma segura
                        resultado['datos']['aspirantes'].append({
                            'id': aspirante.id,
                            'nombre': f"{getattr(aspirante, 'asp_nombre', '')} {getattr(aspirante, 'asp_apellido', '')}",
                            'email': getattr(aspirante, 'asp_correo', ''),
                            'fecha_registro': fecha_registro,
                            'ciudad': getattr(aspirante, 'asp_ciudad', ''),
                            'ocupacion': getattr(aspirante, 'asp_ocupacion', ''),
                            'postulaciones': postulaciones_count,
                            'creditos_turbo': getattr(aspirante, 'asp_creditos_turbo_disponibles', 0)
                        })
                    except Exception as e:
                        print(f"Error procesando aspirante {aspirante.id}: {str(e)}")
                        continue
            
            # Reporte de empresas
            if not tipo_usuario or tipo_usuario == 'empresa':
                empresas_query = Empresa.objects.all()
                
                # Aplicar filtros de fecha si existen los campos
                try:
                    if fecha_inicio and hasattr(Empresa._meta.get_field('em_fecha_registro'), 'name'):
                        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                        empresas_query = empresas_query.filter(em_fecha_registro__gte=fecha_inicio_dt)
                    
                    if fecha_fin and hasattr(Empresa._meta.get_field('em_fecha_registro'), 'name'):
                        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                        empresas_query = empresas_query.filter(em_fecha_registro__lte=fecha_fin_dt)
                except:
                    # Si no existe el campo de fecha, continuar sin filtro
                    pass
                
                for empresa in empresas_query:
                    try:
                        vacantes_count = Vacante.objects.filter(va_idEmpresa_fk=empresa).count()
                        
                        # Obtener fecha de registro de forma segura
                        fecha_registro = None
                        if hasattr(empresa, 'em_fecha_registro') and empresa.em_fecha_registro:
                            fecha_registro = empresa.em_fecha_registro.strftime('%Y-%m-%d %H:%M')
                        
                        resultado['datos']['empresas'].append({
                            'id': empresa.id,
                            'nombre': getattr(empresa, 'em_nombre', ''),
                            'email': getattr(empresa, 'em_email', ''),
                            'nit': getattr(empresa, 'em_nit', ''),
                            'fecha_registro': fecha_registro,
                            'ciudad': getattr(empresa, 'em_ciudad', ''),
                            'sector': getattr(empresa, 'em_sector', ''),
                            'vacantes_publicadas': vacantes_count,
                            'score_turbo': getattr(empresa, 'em_score_turbo', 0)
                        })
                    except Exception as e:
                        print(f"Error procesando empresa {empresa.id}: {str(e)}")
                        continue
            
            # Estadísticas generales
            resultado['resumen']['total_aspirantes'] = len(resultado['datos']['aspirantes'])
            resultado['resumen']['total_empresas'] = len(resultado['datos']['empresas'])
            
            if formato == 'csv':
                # Combinar datos para CSV
                datos_csv = resultado['datos']['aspirantes'] + resultado['datos']['empresas']
                return self._export_csv(datos_csv, 'reporte_usuarios')
            elif formato == 'excel':
                datos_excel = resultado['datos']['aspirantes'] + resultado['datos']['empresas']
                return self._export_excel(datos_excel, 'reporte_usuarios', 'Usuarios')
            elif formato == 'pdf':
                datos_pdf = resultado['datos']['aspirantes'] + resultado['datos']['empresas']
                return self._export_pdf(datos_pdf, 'reporte_usuarios', 'Reporte de Usuarios')
            
            return Response(resultado)
            
        except Exception as e:
            return Response(
                {'error': f'Error generando reporte: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Estadísticas generales para dashboard de reportes"""
        try:
            # Fechas por defecto (30 días)
            fecha_fin = timezone.now()
            fecha_inicio = fecha_fin - timedelta(days=30)
            
            # Parámetros personalizados
            if request.query_params.get('fecha_inicio'):
                fecha_inicio = datetime.strptime(request.query_params.get('fecha_inicio'), '%Y-%m-%d')
            if request.query_params.get('fecha_fin'):
                fecha_fin = datetime.strptime(request.query_params.get('fecha_fin'), '%Y-%m-%d')
            
            # Estadísticas generales
            stats = {
                'periodo': {
                    'inicio': fecha_inicio.strftime('%Y-%m-%d'),
                    'fin': fecha_fin.strftime('%Y-%m-%d')
                },
                'postulaciones': {
                    'total': Postulacion.objects.filter(
                        pos_fechaPostulacion__range=[fecha_inicio, fecha_fin]
                    ).count(),
                    'turbo': Postulacion.objects.filter(
                        pos_fechaPostulacion__range=[fecha_inicio, fecha_fin],
                        pos_es_turbo=True
                    ).count(),
                    'por_estado': list(Postulacion.objects.filter(
                        pos_fechaPostulacion__range=[fecha_inicio, fecha_fin]
                    ).values('pos_estado').annotate(count=Count('id')))
                },
                'vacantes': {
                    'total': Vacante.objects.filter(
                        va_fecha_publicacion__range=[fecha_inicio, fecha_fin]
                    ).count(),
                    'activas': Vacante.objects.filter(
                        va_fecha_publicacion__range=[fecha_inicio, fecha_fin],
                        va_estado='Activa'
                    ).count(),
                    'turbo': Vacante.objects.filter(
                        va_fecha_publicacion__range=[fecha_inicio, fecha_fin],
                        va_modo_turbo=True
                    ).count()
                },
                'usuarios': {
                    'aspirantes_nuevos': Aspirante.objects.filter(
                        asp_fecha_registro__range=[fecha_inicio, fecha_fin]
                    ).count() if hasattr(Aspirante.objects.first(), 'asp_fecha_registro') else 0,
                    'empresas_nuevas': Empresa.objects.filter(
                        em_fecha_registro__range=[fecha_inicio, fecha_fin]
                    ).count() if hasattr(Empresa.objects.first(), 'em_fecha_registro') else 0
                }
            }
            
            return Response(stats)
            
        except Exception as e:
            return Response(
                {'error': f'Error obteniendo estadísticas: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _export_csv(self, data, filename):
        """Exportar datos a CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        
        if not data:
            return response
        
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    def _export_excel(self, data, filename, sheet_name="Reporte"):
        """Exportar datos a Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            wb.save(response)
            return response
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response
    
    def _export_pdf(self, data, filename, title="Reporte de TurboEmpleo"):
        """Exportar datos a PDF con diseño profesional"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        # Configuración de página con márgenes
        doc = SimpleDocTemplate(
            response, 
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=60,
            bottomMargin=60
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Colores corporativos de TurboEmpleo
        turbo_purple = colors.Color(0.369, 0.090, 0.922)  # #5e17eb
        turbo_light_purple = colors.Color(0.839, 0.776, 0.976)  # #d6c6f9
        turbo_dark = colors.Color(0.298, 0.051, 0.800)  # #4c0dcd
        
        # === HEADER CON LOGO Y EMPRESA ===
        header_data = [
            ['', 'TURBOEMPLEO', ''],  # Espacio para logo, nombre, espacio
        ]
        
        header_table = Table(header_data, colWidths=[1.5*inch, 4*inch, 1.5*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), turbo_purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 24),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, 0), 20),
            ('RIGHTPADDING', (0, 0), (-1, 0), 20),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # === TÍTULO DEL REPORTE ===
        title_style = ParagraphStyle(
            'TurboTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            spaceAfter=10,
            spaceBefore=10,
            alignment=1,  # Centrado
            textColor=turbo_dark
        )
        elements.append(Paragraph(title, title_style))
        
        # === INFORMACIÓN DEL REPORTE ===
        info_style = ParagraphStyle(
            'TurboInfo',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            alignment=1,  # Centrado
            textColor=colors.grey,
            spaceAfter=20
        )
        
        fecha_generacion = timezone.now().strftime('%d de %B de %Y a las %H:%M')
        info_text = f"📅 Generado el {fecha_generacion}<br/>🖥️ Sistema de Reportes TurboEmpleo"
        elements.append(Paragraph(info_text, info_style))
        
        # === LÍNEA DECORATIVA ===
        line_data = [[''] * 5]
        line_table = Table(line_data, colWidths=[1.4*inch]*5)
        line_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), turbo_light_purple),
            ('TOPPADDING', (0, 0), (-1, 0), 3),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 25))
        
        if not data:
            # === SIN DATOS ===
            no_data_style = ParagraphStyle(
                'NoData',
                parent=styles['Normal'],
                fontName='Helvetica-Oblique',
                fontSize=14,
                alignment=1,
                textColor=colors.grey,
                spaceAfter=20,
                spaceBefore=20
            )
            elements.append(Paragraph("📋 No hay datos para mostrar en este reporte", no_data_style))
        else:
            # === RESUMEN EJECUTIVO ===
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=12,
                alignment=0,
                textColor=turbo_dark,
                spaceAfter=15
            )
            
            total_registros = len(data)
            summary_text = f"📊 <b>Resumen:</b> Este reporte contiene {total_registros} registro{'s' if total_registros != 1 else ''}"
            elements.append(Paragraph(summary_text, summary_style))
            
            # === TABLA DE DATOS ===
            headers = [header.replace('_', ' ').title() for header in data[0].keys()]
            table_data = [headers]
            
            # Limitar a 50 registros para evitar PDFs muy largos
            limited_data = data[:50]
            for row in limited_data:
                formatted_row = []
                for value in row.values():
                    if isinstance(value, bool):
                        formatted_row.append('✓ Sí' if value else '✗ No')
                    elif isinstance(value, str) and '2025' in str(value):
                        try:
                            # Formatear fechas
                            from datetime import datetime
                            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            formatted_row.append(dt.strftime('%d/%m/%Y %H:%M'))
                        except:
                            formatted_row.append(str(value))
                    else:
                        formatted_row.append(str(value) if value is not None else '-')
                table_data.append(formatted_row)
            
            # Calcular anchos de columna dinámicamente
            num_cols = len(headers)
            if num_cols <= 3:
                col_widths = [2.3*inch] * num_cols
            elif num_cols <= 5:
                col_widths = [1.4*inch] * num_cols
            elif num_cols <= 7:
                col_widths = [1*inch] * num_cols
            else:
                col_widths = [0.8*inch] * num_cols
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # === ESTILO DE TABLA PROFESIONAL ===
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), turbo_purple),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                
                # Bordes y alternancia de colores
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, turbo_light_purple]),
                
                # Línea gruesa bajo header
                ('LINEBELOW', (0, 0), (-1, 0), 2, turbo_dark),
            ]))
            
            elements.append(table)
            
            # === NOTA SI HAY MÁS DATOS ===
            if total_registros > 50:
                note_style = ParagraphStyle(
                    'Note',
                    parent=styles['Normal'],
                    fontName='Helvetica-Oblique',
                    fontSize=10,
                    alignment=1,
                    textColor=colors.grey,
                    spaceAfter=10,
                    spaceBefore=15
                )
                note_text = f"📝 <i>Nota: Se muestran los primeros 50 de {total_registros} registros. Para ver todos los datos, use la exportación en Excel.</i>"
                elements.append(Paragraph(note_text, note_style))
        
        # === FOOTER ===
        elements.append(Spacer(1, 30))
        footer_data = [['TurboEmpleo - Sistema de Gestión de Empleo | www.turboempleo.com']]
        footer_table = Table(footer_data, colWidths=[7*inch])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), turbo_light_purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), turbo_dark),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ]))
        elements.append(footer_table)
        
        # Construir PDF
        doc.build(elements)
        return response